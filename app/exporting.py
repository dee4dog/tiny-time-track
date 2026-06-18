"""Excel (.xlsx) exports - the accountant's escape hatch.

Builds workbooks in memory with openpyxl and returns the bytes. Money is
written as real numbers (not pre-formatted strings) with a ZAR number format
so the accountant can sum/filter in Excel.
"""
from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app import costing
from app.dashboard_service import people_overview
from app.models import Project, ProjectStatus
from app.settings_store import get_setting

_HEADER_FILL = PatternFill("solid", fgColor="2F6FED")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _money_format(db: Session) -> str:
    symbol = get_setting(db, "currency_symbol") or "R"
    # e.g.  R #,##0  with a space thousands separator rendered by Excel locale
    return f'"{symbol} "#,##0;[Red]("{symbol} "#,##0)'


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left")
    ws.freeze_panes = "A2"


def _autofit(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_projects_xlsx(db: Session) -> bytes:
    money_fmt = _money_format(db)
    projects = list(
        db.scalars(
            select(Project)
            .where(Project.status != ProjectStatus.archived)
            .order_by(Project.billable.desc(), Project.name)
        )
    )
    summaries = costing.summarise_all_projects(db, projects)

    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"
    headers = ["Project", "Number", "Billable", "Fee", "Hours", "Cost to date", "Profit", "Margin %"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for s in summaries:
        row = [
            s.project.display_name,
            s.project.number,
            "Yes" if s.billable else "No",
            float(s.fee) if s.billable else None,
            float(s.hours),
            float(s.cost),
            float(s.profit) if s.profit is not None else None,
            float(s.margin) if s.margin is not None else None,
        ]
        ws.append(row)
        r = ws.max_row
        for c in (4, 6, 7):
            ws.cell(row=r, column=c).number_format = money_fmt
        ws.cell(row=r, column=5).number_format = "0.0"
        ws.cell(row=r, column=8).number_format = "0%"

    _autofit(ws, [34, 12, 10, 16, 10, 16, 16, 10])
    ws.cell(row=ws.max_row + 2, column=1, value=f"Generated {date.today().isoformat()}")
    return _to_bytes(wb)


def export_people_xlsx(db: Session) -> bytes:
    stats = people_overview(db)
    money_fmt = _money_format(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "People"
    headers = ["Employee", "Role", "Capacity (hrs)", "Planned", "Actual", "Overtime", "Utilisation %", "Cost (YTD)"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for s in stats:
        ws.append([
            s.employee.name,
            s.employee.role.value,
            s.capacity,
            float(s.planned),
            float(s.actual),
            float(s.overtime),
            float(s.utilisation) if s.utilisation is not None else None,
            float(s.cost),
        ])
        r = ws.max_row
        for c in (4, 5, 6):
            ws.cell(row=r, column=c).number_format = "0.0"
        ws.cell(row=r, column=7).number_format = "0%"
        ws.cell(row=r, column=8).number_format = money_fmt

    _autofit(ws, [22, 12, 14, 12, 12, 12, 14, 16])
    ws.cell(row=ws.max_row + 2, column=1, value=f"Generated {date.today().isoformat()}")
    return _to_bytes(wb)
